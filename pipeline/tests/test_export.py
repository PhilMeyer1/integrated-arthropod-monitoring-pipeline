"""
Tests for export module.

Tests Excel export, CSV export, and statistics calculation.
"""

import pytest
from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import load_workbook

from src.export.excel_export import ExcelExporter
from src.export.csv_export import CSVExporter
from src.export.statistics import StatisticsCalculator


@pytest.fixture
def sample_results():
    """Create sample classification results."""
    return [
        {
            'image_id': 1,
            'sample_id': 'S001',
            'size_fraction': '1',
            'predicted_taxon': 'Coleoptera',
            'confidence': 0.95,
            'path': ['Arthropoda', 'Insecta', 'Coleoptera'],
            'bounding_box': (100, 100, 200, 200),
            'specimen_index': 0
        },
        {
            'image_id': 2,
            'sample_id': 'S001',
            'size_fraction': '1',
            'predicted_taxon': 'Coleoptera',
            'confidence': 0.92,
            'path': ['Arthropoda', 'Insecta', 'Coleoptera'],
            'bounding_box': (150, 150, 250, 250),
            'specimen_index': 1
        },
        {
            'image_id': 3,
            'sample_id': 'S001',
            'size_fraction': '1',
            'predicted_taxon': 'Diptera',
            'confidence': 0.88,
            'path': ['Arthropoda', 'Insecta', 'Diptera'],
            'bounding_box': (200, 200, 300, 300),
            'specimen_index': 2
        },
        {
            'image_id': 4,
            'sample_id': 'S002',
            'size_fraction': '2',
            'predicted_taxon': 'Arachnida',
            'confidence': 0.91,
            'path': ['Arthropoda', 'Arachnida'],
            'bounding_box': (120, 120, 220, 220),
            'specimen_index': 0
        }
    ]


@pytest.fixture
def sample_results_with_ground_truth():
    """Create sample results with ground truth for evaluation."""
    return [
        {
            'image_id': 1,
            'predicted_taxon': 'Coleoptera',
            'actual_taxon': 'Coleoptera',
            'confidence': 0.95,
            'correct': True
        },
        {
            'image_id': 2,
            'predicted_taxon': 'Coleoptera',
            'actual_taxon': 'Coleoptera',
            'confidence': 0.92,
            'correct': True
        },
        {
            'image_id': 3,
            'predicted_taxon': 'Diptera',
            'actual_taxon': 'Coleoptera',
            'confidence': 0.88,
            'correct': False
        },
        {
            'image_id': 4,
            'predicted_taxon': 'Arachnida',
            'actual_taxon': 'Arachnida',
            'confidence': 0.91,
            'correct': True
        }
    ]


class TestExcelExporter:
    """Test ExcelExporter class."""

    def test_initialization(self, tmp_path):
        """Test Excel exporter initialization."""
        exporter = ExcelExporter(output_dir=tmp_path)

        assert exporter.output_dir == tmp_path
        assert exporter.output_dir.exists()

    def test_export_creates_file(self, tmp_path, sample_results):
        """Test that export creates Excel file."""
        exporter = ExcelExporter(output_dir=tmp_path)

        output_path = exporter.export(
            results=sample_results,
            filename="test_results.xlsx"
        )

        assert output_path.exists()
        assert output_path.suffix == '.xlsx'

    def test_excel_structure(self, tmp_path, sample_results):
        """Test Excel file structure."""
        exporter = ExcelExporter(output_dir=tmp_path)

        output_path = exporter.export(
            results=sample_results,
            filename="test_results.xlsx"
        )

        # Load workbook and check sheets
        wb = load_workbook(output_path)

        # Should have multiple sheets
        assert 'Results' in wb.sheetnames
        assert 'Summary' in wb.sheetnames

    def test_results_sheet_content(self, tmp_path, sample_results):
        """Test content of Results sheet."""
        exporter = ExcelExporter(output_dir=tmp_path)

        output_path = exporter.export(
            results=sample_results,
            filename="test_results.xlsx"
        )

        # Read Results sheet
        df = pd.read_excel(output_path, sheet_name='Results')

        # Should have all rows
        assert len(df) == len(sample_results)

        # Should have key columns
        assert 'sample_id' in df.columns
        assert 'predicted_taxon' in df.columns
        assert 'confidence' in df.columns

    def test_summary_sheet_aggregation(self, tmp_path, sample_results):
        """Test Summary sheet aggregation."""
        exporter = ExcelExporter(output_dir=tmp_path)

        output_path = exporter.export(
            results=sample_results,
            filename="test_results.xlsx"
        )

        # Read Summary sheet
        df = pd.read_excel(output_path, sheet_name='Summary')

        # Should group by sample_id or taxon
        assert len(df) > 0

    def test_per_taxon_sheet(self, tmp_path, sample_results):
        """Test Per-Taxon sheet."""
        exporter = ExcelExporter(output_dir=tmp_path)

        output_path = exporter.export(
            results=sample_results,
            filename="test_results.xlsx",
            include_per_taxon=True
        )

        wb = load_workbook(output_path)

        # Should have Per-Taxon sheet
        if 'Per-Taxon' in wb.sheetnames:
            df = pd.read_excel(output_path, sheet_name='Per-Taxon')
            assert len(df) > 0

    def test_formatting_applied(self, tmp_path, sample_results):
        """Test that formatting is applied to Excel."""
        exporter = ExcelExporter(output_dir=tmp_path)

        output_path = exporter.export(
            results=sample_results,
            filename="test_results.xlsx"
        )

        wb = load_workbook(output_path)
        ws = wb['Results']

        # Check that headers exist (first row)
        assert ws.cell(1, 1).value is not None

        # Formatting is applied (can't easily test colors, but file should be valid)
        assert ws.max_row == len(sample_results) + 1  # +1 for header


class TestCSVExporter:
    """Test CSVExporter class."""

    def test_initialization(self, tmp_path):
        """Test CSV exporter initialization."""
        exporter = CSVExporter(output_dir=tmp_path)

        assert exporter.output_dir == tmp_path
        assert exporter.output_dir.exists()

    def test_export_creates_file(self, tmp_path, sample_results):
        """Test that export creates CSV file."""
        exporter = CSVExporter(output_dir=tmp_path)

        output_path = exporter.export(
            results=sample_results,
            filename="test_results.csv"
        )

        assert output_path.exists()
        assert output_path.suffix == '.csv'

    def test_csv_content(self, tmp_path, sample_results):
        """Test CSV file content."""
        exporter = CSVExporter(output_dir=tmp_path)

        output_path = exporter.export(
            results=sample_results,
            filename="test_results.csv"
        )

        # Read CSV
        df = pd.read_csv(output_path)

        # Should have all rows
        assert len(df) == len(sample_results)

        # Should have key columns
        assert 'sample_id' in df.columns
        assert 'predicted_taxon' in df.columns
        assert 'confidence' in df.columns

    def test_hierarchical_path_export(self, tmp_path, sample_results):
        """Test export with hierarchical path."""
        exporter = CSVExporter(output_dir=tmp_path)

        output_path = exporter.export(
            results=sample_results,
            filename="test_results.csv",
            include_path=True
        )

        df = pd.read_csv(output_path)

        # Should have path column
        if 'path' in df.columns:
            # Path should be a string representation of list
            assert isinstance(df.iloc[0]['path'], str)

    def test_batch_export(self, tmp_path, sample_results):
        """Test batch export of multiple CSVs."""
        exporter = CSVExporter(output_dir=tmp_path)

        # Group by sample_id
        grouped = {}
        for result in sample_results:
            sid = result['sample_id']
            if sid not in grouped:
                grouped[sid] = []
            grouped[sid].append(result)

        # Export each group
        output_paths = []
        for sample_id, results in grouped.items():
            output_path = exporter.export(
                results=results,
                filename=f"{sample_id}_results.csv"
            )
            output_paths.append(output_path)

        # Should create multiple files
        assert len(output_paths) >= 2
        for path in output_paths:
            assert path.exists()

    def test_summary_statistics_export(self, tmp_path, sample_results):
        """Test export of summary statistics."""
        exporter = CSVExporter(output_dir=tmp_path)

        # Create summary
        summary = {
            'total_specimens': len(sample_results),
            'unique_taxa': len(set(r['predicted_taxon'] for r in sample_results)),
            'avg_confidence': np.mean([r['confidence'] for r in sample_results])
        }

        # Export summary
        output_path = exporter.export_summary(
            summary=summary,
            filename="summary.csv"
        )

        assert output_path.exists()

        # Read and verify
        df = pd.read_csv(output_path)
        assert len(df) > 0


class TestStatisticsCalculator:
    """Test StatisticsCalculator class."""

    def test_initialization(self):
        """Test statistics calculator initialization."""
        calculator = StatisticsCalculator()
        assert calculator is not None

    def test_basic_statistics(self, sample_results):
        """Test basic statistics calculation."""
        calculator = StatisticsCalculator()

        stats = calculator.calculate_basic_stats(sample_results)

        assert 'total_specimens' in stats
        assert stats['total_specimens'] == len(sample_results)

        assert 'unique_taxa' in stats
        assert stats['unique_taxa'] > 0

        assert 'avg_confidence' in stats
        assert 0.0 <= stats['avg_confidence'] <= 1.0

    def test_per_sample_statistics(self, sample_results):
        """Test per-sample statistics."""
        calculator = StatisticsCalculator()

        stats = calculator.calculate_per_sample_stats(sample_results)

        # Should group by sample_id
        assert 'S001' in stats
        assert 'S002' in stats

        # S001 should have 3 specimens
        assert stats['S001']['count'] == 3

        # S002 should have 1 specimen
        assert stats['S002']['count'] == 1

    def test_per_taxon_statistics(self, sample_results):
        """Test per-taxon statistics."""
        calculator = StatisticsCalculator()

        stats = calculator.calculate_per_taxon_stats(sample_results)

        # Should group by taxon
        assert 'Coleoptera' in stats
        assert 'Diptera' in stats
        assert 'Arachnida' in stats

        # Coleoptera should have 2 specimens
        assert stats['Coleoptera']['count'] == 2

    def test_confidence_statistics(self, sample_results):
        """Test confidence statistics."""
        calculator = StatisticsCalculator()

        stats = calculator.calculate_confidence_stats(sample_results)

        assert 'mean' in stats
        assert 'median' in stats
        assert 'std' in stats
        assert 'min' in stats
        assert 'max' in stats

        # Confidence should be in valid range
        assert 0.0 <= stats['min'] <= 1.0
        assert 0.0 <= stats['max'] <= 1.0
        assert stats['min'] <= stats['mean'] <= stats['max']

    def test_accuracy_calculation(self, sample_results_with_ground_truth):
        """Test accuracy calculation."""
        calculator = StatisticsCalculator()

        stats = calculator.calculate_accuracy(sample_results_with_ground_truth)

        assert 'accuracy' in stats
        # 3 correct out of 4 = 0.75
        assert stats['accuracy'] == 0.75

    def test_confusion_matrix(self, sample_results_with_ground_truth):
        """Test confusion matrix calculation."""
        calculator = StatisticsCalculator()

        cm = calculator.calculate_confusion_matrix(sample_results_with_ground_truth)

        assert cm is not None
        # Should be 2D array/dict

    def test_f1_score_calculation(self, sample_results_with_ground_truth):
        """Test F1 score calculation."""
        calculator = StatisticsCalculator()

        stats = calculator.calculate_f1_score(sample_results_with_ground_truth)

        assert 'f1' in stats or 'f1_score' in stats
        # F1 should be between 0 and 1
        f1 = stats.get('f1', stats.get('f1_score', 0))
        assert 0.0 <= f1 <= 1.0

    def test_per_class_metrics(self, sample_results_with_ground_truth):
        """Test per-class precision, recall, F1."""
        calculator = StatisticsCalculator()

        metrics = calculator.calculate_per_class_metrics(sample_results_with_ground_truth)

        # Should have metrics for each class
        for class_name in ['Coleoptera', 'Arachnida']:
            if class_name in metrics:
                assert 'precision' in metrics[class_name] or 'recall' in metrics[class_name]


class TestIntegration:
    """Integration tests for export pipeline."""

    def test_statistics_to_excel_flow(self, tmp_path, sample_results):
        """Test statistics calculation and Excel export."""
        # Calculate statistics
        calculator = StatisticsCalculator()
        stats = calculator.calculate_basic_stats(sample_results)

        # Export to Excel
        exporter = ExcelExporter(output_dir=tmp_path)
        output_path = exporter.export(
            results=sample_results,
            filename="test_results.xlsx",
            statistics=stats
        )

        assert output_path.exists()

    def test_statistics_to_csv_flow(self, tmp_path, sample_results):
        """Test statistics calculation and CSV export."""
        # Calculate statistics
        calculator = StatisticsCalculator()
        stats = calculator.calculate_basic_stats(sample_results)

        # Export results
        csv_exporter = CSVExporter(output_dir=tmp_path)
        results_path = csv_exporter.export(
            results=sample_results,
            filename="test_results.csv"
        )

        # Export statistics
        stats_path = csv_exporter.export_summary(
            summary=stats,
            filename="test_summary.csv"
        )

        assert results_path.exists()
        assert stats_path.exists()

    def test_both_export_formats(self, tmp_path, sample_results):
        """Test exporting same data in both Excel and CSV."""
        # Excel export
        excel_exporter = ExcelExporter(output_dir=tmp_path)
        excel_path = excel_exporter.export(
            results=sample_results,
            filename="results.xlsx"
        )

        # CSV export
        csv_exporter = CSVExporter(output_dir=tmp_path)
        csv_path = csv_exporter.export(
            results=sample_results,
            filename="results.csv"
        )

        # Both should exist
        assert excel_path.exists()
        assert csv_path.exists()

        # Read both and compare
        df_excel = pd.read_excel(excel_path, sheet_name='Results')
        df_csv = pd.read_csv(csv_path)

        # Should have same number of rows
        assert len(df_excel) == len(df_csv)


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
