"""
Excel export functionality for classification results.

This module exports classification results, taxonomy hierarchies, and
statistics to Excel format with proper formatting and multiple sheets.

Corresponds to Methods section "Data Export and Analysis".
"""

from pathlib import Path
from typing import Dict, List, Optional
import json

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from src.utils.logging_config import get_logger

logger = get_logger(__name__)


class ExcelExporter:
    """
    Export classification results and metadata to Excel.

    Creates multi-sheet Excel workbooks with:
    - Classification results per specimen
    - Taxonomy hierarchy
    - Statistics per taxon
    - Model performance metrics
    - Confusion matrices

    Example:
        >>> exporter = ExcelExporter()
        >>> exporter.export_classification_results(
        ...     results_df=results,
        ...     output_path=Path('./results.xlsx'),
        ...     include_hierarchy=True
        ... )
    """

    def __init__(self, data_manager=None):
        """
        Initialize Excel exporter.

        Args:
            data_manager: DataManager for database access
        """
        self.data_manager = data_manager

    def apply_header_formatting(self, worksheet, num_columns: int):
        """
        Apply formatting to header row.

        Args:
            worksheet: openpyxl worksheet
            num_columns: Number of columns to format
        """
        # Header styling
        header_fill = PatternFill(
            start_color='366092',
            end_color='366092',
            fill_type='solid'
        )
        header_font = Font(
            name='Arial',
            size=11,
            bold=True,
            color='FFFFFF'
        )

        for col in range(1, num_columns + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Freeze header row
        worksheet.freeze_panes = 'A2'

    def auto_adjust_column_widths(self, worksheet):
        """
        Auto-adjust column widths based on content.

        Args:
            worksheet: openpyxl worksheet
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Max 50 chars
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def export_classification_results(
        self,
        results_df: pd.DataFrame,
        output_path: Path,
        include_hierarchy: bool = True,
        include_statistics: bool = True
    ):
        """
        Export classification results to Excel.

        Args:
            results_df: DataFrame with classification results
            output_path: Path to save Excel file
            include_hierarchy: Include taxonomy hierarchy sheet
            include_statistics: Include statistics sheet

        Example:
            >>> exporter.export_classification_results(
            ...     results_df,
            ...     Path('./classification_results.xlsx')
            ... )
        """
        logger.info(f"Exporting classification results to {output_path}")

        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: Classification Results
            results_df.to_excel(
                writer,
                sheet_name='Classification Results',
                index=False
            )

            # Format results sheet
            worksheet = writer.sheets['Classification Results']
            self.apply_header_formatting(worksheet, len(results_df.columns))
            self.auto_adjust_column_widths(worksheet)

            # Sheet 2: Summary Statistics
            if include_statistics:
                stats_df = self.calculate_summary_statistics(results_df)
                stats_df.to_excel(
                    writer,
                    sheet_name='Summary Statistics',
                    index=False
                )

                worksheet = writer.sheets['Summary Statistics']
                self.apply_header_formatting(worksheet, len(stats_df.columns))
                self.auto_adjust_column_widths(worksheet)

            # Sheet 3: Per-Taxon Statistics
            if include_statistics:
                taxon_stats_df = self.calculate_per_taxon_statistics(results_df)
                taxon_stats_df.to_excel(
                    writer,
                    sheet_name='Per-Taxon Statistics',
                    index=False
                )

                worksheet = writer.sheets['Per-Taxon Statistics']
                self.apply_header_formatting(worksheet, len(taxon_stats_df.columns))
                self.auto_adjust_column_widths(worksheet)

            # Sheet 4: Taxonomy Hierarchy
            if include_hierarchy and self.data_manager:
                hierarchy_df = self.build_hierarchy_dataframe()
                if not hierarchy_df.empty:
                    hierarchy_df.to_excel(
                        writer,
                        sheet_name='Taxonomy Hierarchy',
                        index=False
                    )

                    worksheet = writer.sheets['Taxonomy Hierarchy']
                    self.apply_header_formatting(worksheet, len(hierarchy_df.columns))
                    self.auto_adjust_column_widths(worksheet)

        logger.info(f"Exported {len(results_df)} results to {output_path}")

    def calculate_summary_statistics(self, results_df: pd.DataFrame) -> pd.DataFrame:
        """
        Calculate overall summary statistics.

        Args:
            results_df: Classification results DataFrame

        Returns:
            DataFrame with summary statistics
        """
        # Calculate metrics
        total_predictions = len(results_df)
        correct_predictions = (
            results_df['predicted_taxon_id'] == results_df['correct_taxon_id']
        ).sum()

        accuracy = correct_predictions / total_predictions if total_predictions > 0 else 0

        avg_confidence = results_df['confidence'].mean()
        median_confidence = results_df['confidence'].median()
        min_confidence = results_df['confidence'].min()
        max_confidence = results_df['confidence'].max()

        # Test set statistics
        test_set = results_df[results_df['is_excluded'] == True]
        test_accuracy = 0
        if len(test_set) > 0:
            test_correct = (
                test_set['predicted_taxon_id'] == test_set['correct_taxon_id']
            ).sum()
            test_accuracy = test_correct / len(test_set)

        # Create summary DataFrame
        summary = pd.DataFrame([
            {'Metric': 'Total Predictions', 'Value': total_predictions},
            {'Metric': 'Correct Predictions', 'Value': correct_predictions},
            {'Metric': 'Overall Accuracy', 'Value': f'{accuracy:.3f}'},
            {'Metric': 'Test Set Size', 'Value': len(test_set)},
            {'Metric': 'Test Set Accuracy', 'Value': f'{test_accuracy:.3f}'},
            {'Metric': 'Average Confidence', 'Value': f'{avg_confidence:.3f}'},
            {'Metric': 'Median Confidence', 'Value': f'{median_confidence:.3f}'},
            {'Metric': 'Min Confidence', 'Value': f'{min_confidence:.3f}'},
            {'Metric': 'Max Confidence', 'Value': f'{max_confidence:.3f}'},
        ])

        return summary

    def calculate_per_taxon_statistics(self, results_df: pd.DataFrame) -> pd.DataFrame:
        """
        Calculate statistics per model/taxon.

        Args:
            results_df: Classification results DataFrame

        Returns:
            DataFrame with per-taxon statistics
        """
        taxon_stats = []

        for model_taxon in results_df['model_taxon_id'].unique():
            taxon_df = results_df[results_df['model_taxon_id'] == model_taxon]

            total = len(taxon_df)
            correct = (
                taxon_df['predicted_taxon_id'] == taxon_df['correct_taxon_id']
            ).sum()
            accuracy = correct / total if total > 0 else 0

            avg_conf = taxon_df['confidence'].mean()

            # Test set for this taxon
            test_set = taxon_df[taxon_df['is_excluded'] == True]
            test_accuracy = 0
            if len(test_set) > 0:
                test_correct = (
                    test_set['predicted_taxon_id'] == test_set['correct_taxon_id']
                ).sum()
                test_accuracy = test_correct / len(test_set)

            taxon_stats.append({
                'Model Taxon': model_taxon,
                'Total Predictions': total,
                'Correct': correct,
                'Accuracy': f'{accuracy:.3f}',
                'Test Set Size': len(test_set),
                'Test Accuracy': f'{test_accuracy:.3f}',
                'Avg Confidence': f'{avg_conf:.3f}'
            })

        return pd.DataFrame(taxon_stats)

    def build_hierarchy_dataframe(self) -> pd.DataFrame:
        """
        Build taxonomy hierarchy as flat DataFrame.

        Returns:
            DataFrame with hierarchy (taxon_id, name, rank, parent_id)
        """
        if self.data_manager is None:
            return pd.DataFrame()

        # Get all taxa
        taxa = self.data_manager.get_all_taxa()

        hierarchy_data = []
        for taxon in taxa:
            hierarchy_data.append({
                'Taxon ID': taxon['taxon_id'],
                'Scientific Name': taxon.get('scientific_name', ''),
                'Rank': taxon.get('rank', ''),
                'Parent ID': taxon.get('parent_id', ''),
                'Common Name': taxon.get('common_name', '')
            })

        return pd.DataFrame(hierarchy_data)

    def export_model_comparison(
        self,
        set_numbers: List[int],
        output_path: Path
    ):
        """
        Export comparison of multiple model sets.

        Args:
            set_numbers: List of set numbers to compare
            output_path: Path to save Excel file

        Example:
            >>> exporter.export_model_comparison([1, 2, 3], Path('./comparison.xlsx'))
        """
        if self.data_manager is None:
            logger.error("Data manager required for model comparison")
            return

        logger.info(f"Exporting comparison for sets {set_numbers}")

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for set_num in set_numbers:
                # Get results for this set
                results = self.data_manager.get_inference_results(set_num)
                results_df = pd.DataFrame(results)

                if results_df.empty:
                    logger.warning(f"No results for set {set_num}")
                    continue

                # Export to sheet
                sheet_name = f'Set {set_num}'
                results_df.to_excel(writer, sheet_name=sheet_name, index=False)

                # Format
                worksheet = writer.sheets[sheet_name]
                self.apply_header_formatting(worksheet, len(results_df.columns))
                self.auto_adjust_column_widths(worksheet)

            # Summary comparison sheet
            comparison_df = self.build_set_comparison(set_numbers)
            comparison_df.to_excel(writer, sheet_name='Comparison', index=False)

            worksheet = writer.sheets['Comparison']
            self.apply_header_formatting(worksheet, len(comparison_df.columns))
            self.auto_adjust_column_widths(worksheet)

        logger.info(f"Exported comparison to {output_path}")

    def build_set_comparison(self, set_numbers: List[int]) -> pd.DataFrame:
        """
        Build comparison table for multiple sets.

        Args:
            set_numbers: List of set numbers

        Returns:
            DataFrame with comparison metrics
        """
        comparison_data = []

        for set_num in set_numbers:
            results = self.data_manager.get_inference_results(set_num)
            results_df = pd.DataFrame(results)

            if results_df.empty:
                continue

            # Calculate metrics
            total = len(results_df)
            correct = (
                results_df['predicted_taxon_id'] == results_df['correct_taxon_id']
            ).sum()
            accuracy = correct / total if total > 0 else 0

            # Test set
            test_set = results_df[results_df['is_excluded'] == True]
            test_accuracy = 0
            if len(test_set) > 0:
                test_correct = (
                    test_set['predicted_taxon_id'] == test_set['correct_taxon_id']
                ).sum()
                test_accuracy = test_correct / len(test_set)

            comparison_data.append({
                'Set Number': set_num,
                'Total Predictions': total,
                'Overall Accuracy': f'{accuracy:.3f}',
                'Test Set Size': len(test_set),
                'Test Accuracy': f'{test_accuracy:.3f}',
                'Avg Confidence': f'{results_df["confidence"].mean():.3f}'
            })

        return pd.DataFrame(comparison_data)

    def export_specimen_details(
        self,
        image_ids: List[int],
        output_path: Path
    ):
        """
        Export detailed information for specific specimens.

        Args:
            image_ids: List of image IDs
            output_path: Path to save Excel file

        Example:
            >>> exporter.export_specimen_details([1, 2, 3], Path('./specimens.xlsx'))
        """
        if self.data_manager is None:
            logger.error("Data manager required")
            return

        logger.info(f"Exporting details for {len(image_ids)} specimens")

        specimen_data = []

        for img_id in image_ids:
            # Get specimen info
            specimen = self.data_manager.get_specimen(img_id)

            if specimen:
                specimen_data.append({
                    'Image ID': img_id,
                    'Sample ID': specimen.get('sample_id', ''),
                    'Size Fraction': specimen.get('size_fraction', ''),
                    'Manual Determination': specimen.get('determination_manual', ''),
                    'File Path': specimen.get('file_path', ''),
                    'Width': specimen.get('width', ''),
                    'Height': specimen.get('height', ''),
                    'Date': specimen.get('date', '')
                })

        df = pd.DataFrame(specimen_data)

        df.to_excel(output_path, index=False, engine='openpyxl')

        # Format
        workbook = openpyxl.load_workbook(output_path)
        worksheet = workbook.active
        self.apply_header_formatting(worksheet, len(df.columns))
        self.auto_adjust_column_widths(worksheet)
        workbook.save(output_path)

        logger.info(f"Exported specimen details to {output_path}")
